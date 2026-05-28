"""
Excel converter (.xlsx and legacy .xls).

Uses openpyxl for .xlsx (structure preservation, merged cells, named sheets).
Uses xlrd for legacy .xls files.
Falls back to pandas read_excel when structure is not critical.

Each worksheet is converted to a Markdown section with a table.
Merged cells are expanded so every row has a consistent column count.
"""

import os
from typing import Optional, Callable

from .confidence import ConfidenceResult
from .markdown_writer import ConversionOutput, rows_to_markdown_table, _pad
from .logger import ConversionLogger


def convert(
    source_file: str,
    alias: str = "",
    logger: Optional[ConversionLogger] = None,
    progress_callback: Optional[Callable[[float], None]] = None,
) -> ConversionOutput:
    ext = os.path.splitext(source_file)[1].lower()
    if ext in (".xls",):
        return _convert_xls(source_file, alias, logger, progress_callback)
    return _convert_xlsx(source_file, alias, logger, progress_callback)


# ---------------------------------------------------------------------------
# .xlsx via openpyxl
# ---------------------------------------------------------------------------

def _convert_xlsx(
    source_file: str,
    alias: str,
    logger: Optional[ConversionLogger],
    progress_callback: Optional[Callable[[float], None]],
) -> ConversionOutput:
    output = ConversionOutput(source_file=source_file, alias=alias)
    confidence = ConfidenceResult(source_file=source_file)
    output.confidence = confidence
    output.engine_used = "openpyxl"

    def log_info(msg):
        if logger: logger.info(msg)
    def log_warn(msg):
        if logger: logger.warning(msg)
        confidence.add_warning(msg)
    def progress(p):
        if progress_callback: progress_callback(p)

    log_info(f"XLSX converter started | file={os.path.basename(source_file)}")
    progress(0.05)

    try:
        import openpyxl
    except ImportError:
        log_warn("openpyxl not installed — falling back to pandas.")
        return _pandas_fallback(source_file, alias, logger, progress_callback)

    try:
        wb = openpyxl.load_workbook(source_file, data_only=True)
    except Exception as e:
        log_warn(f"openpyxl failed to open file: {e}")
        return _pandas_fallback(source_file, alias, logger, progress_callback)

    sheets = wb.sheetnames
    log_info(f"Opened workbook | sheets={len(sheets)}")
    progress(0.1)

    table_scores = []
    text_scores = []
    any_data = False

    for sheet_idx, sheet_name in enumerate(sheets):
        ws = wb[sheet_name]
        sheet_progress_start = 0.1 + (sheet_idx / len(sheets)) * 0.8
        sheet_progress_end = 0.1 + ((sheet_idx + 1) / len(sheets)) * 0.8

        # Expand merged cells — fill each cell in the merge range with the top-left value
        merge_map = {}
        for merge_range in ws.merged_cells.ranges:
            top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    merge_map[(row, col)] = top_left_value

        all_rows = []
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                if (cell.row, cell.column) in merge_map:
                    val = merge_map[(cell.row, cell.column)]
                else:
                    val = cell.value
                cells.append("" if val is None else str(val))
            # Skip entirely empty rows
            if any(c.strip() for c in cells):
                all_rows.append(cells)

        progress(sheet_progress_start + (sheet_progress_end - sheet_progress_start) * 0.5)

        if not all_rows:
            log_info(f"Sheet '{sheet_name}' is empty — skipped.")
            continue

        any_data = True
        headers = all_rows[0]
        rows = all_rows[1:]

        # Normalize column count
        col_count = max(len(headers), max((len(r) for r in rows), default=0))
        headers = _pad(headers, col_count)
        rows = [_pad(r, col_count) for r in rows]

        table_md = rows_to_markdown_table(headers, rows)
        meta = (
            f"*Sheet: **{sheet_name}**  "
            f"Rows: {len(rows)}  Columns: {len(headers)}*"
        )
        body = f"{meta}\n\n{table_md}"
        heading = f"## {sheet_name}"
        output.add_section(body=body, heading=heading)

        # Confidence per sheet
        ragged = sum(1 for r in rows if len(r) != col_count)
        if ragged == 0:
            table_scores.append("High")
        elif ragged / max(len(rows), 1) < 0.1:
            table_scores.append("Medium")
        else:
            table_scores.append("Low")
            log_warn(f"Sheet '{sheet_name}': {ragged} ragged rows detected.")

        text_scores.append("High")
        log_info(f"Sheet '{sheet_name}' converted | rows={len(rows)} cols={col_count}")
        progress(sheet_progress_end)

    if not any_data:
        log_warn("Workbook contained no data.")
        confidence.text_extraction = "Low"
        confidence.table_structure = "Low"
    else:
        priority = {"High": 3, "Medium": 2, "Low": 1, "Failed": 0}
        confidence.table_structure = min(table_scores, key=lambda s: priority.get(s, 0)) if table_scores else "N/A"
        confidence.text_extraction = "High"

    confidence.image_extraction = "N/A"
    confidence.image_placement = "N/A"
    confidence.ocr_confidence = "N/A"
    confidence.document_order = "High"
    confidence.derive_overall()

    progress(1.0)
    log_info("XLSX conversion complete.")
    return output


# ---------------------------------------------------------------------------
# Legacy .xls via xlrd
# ---------------------------------------------------------------------------

def _convert_xls(
    source_file: str,
    alias: str,
    logger: Optional[ConversionLogger],
    progress_callback: Optional[Callable[[float], None]],
) -> ConversionOutput:
    output = ConversionOutput(source_file=source_file, alias=alias)
    confidence = ConfidenceResult(source_file=source_file)
    output.confidence = confidence
    output.engine_used = "xlrd"

    def log_info(msg):
        if logger: logger.info(msg)
    def log_warn(msg):
        if logger: logger.warning(msg)
        confidence.add_warning(msg)
    def progress(p):
        if progress_callback: progress_callback(p)

    log_info(f"XLS converter started | file={os.path.basename(source_file)}")
    progress(0.05)

    try:
        import xlrd
    except ImportError:
        log_warn("xlrd not installed — falling back to pandas.")
        return _pandas_fallback(source_file, alias, logger, progress_callback)

    try:
        wb = xlrd.open_workbook(source_file)
    except Exception as e:
        log_warn(f"xlrd failed: {e}")
        return _pandas_fallback(source_file, alias, logger, progress_callback)

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if ws.nrows == 0:
            continue
        all_rows = []
        for r in range(ws.nrows):
            cells = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
            if any(c.strip() for c in cells):
                all_rows.append(cells)

        if not all_rows:
            continue

        headers = all_rows[0]
        rows = all_rows[1:]
        col_count = max(len(headers), max((len(r) for r in rows), default=0))
        headers = _pad(headers, col_count)
        rows = [_pad(r, col_count) for r in rows]

        table_md = rows_to_markdown_table(headers, rows)
        meta = f"*Sheet: **{ws.name}**  Rows: {len(rows)}  Columns: {len(headers)}*"
        output.add_section(body=f"{meta}\n\n{table_md}", heading=f"## {ws.name}")
        log_info(f"Sheet '{ws.name}' converted | rows={len(rows)}")
        progress(0.1 + (sheet_idx + 1) / wb.nsheets * 0.8)

    confidence.text_extraction = "High"
    confidence.table_structure = "High"
    confidence.image_extraction = "N/A"
    confidence.image_placement = "N/A"
    confidence.ocr_confidence = "N/A"
    confidence.document_order = "High"
    confidence.derive_overall()
    progress(1.0)
    log_info("XLS conversion complete.")
    return output


# ---------------------------------------------------------------------------
# pandas fallback (handles both .xlsx and .xls)
# ---------------------------------------------------------------------------

def _pandas_fallback(
    source_file: str,
    alias: str,
    logger: Optional[ConversionLogger],
    progress_callback: Optional[Callable[[float], None]],
) -> ConversionOutput:
    output = ConversionOutput(source_file=source_file, alias=alias)
    confidence = ConfidenceResult(source_file=source_file)
    output.confidence = confidence
    output.engine_used = "pandas_excel"

    def log_info(msg):
        if logger: logger.info(msg)
    def log_warn(msg):
        if logger: logger.warning(msg)

    log_info("Using pandas Excel fallback.")

    try:
        import pandas as pd
        sheets = pd.read_excel(source_file, sheet_name=None, dtype=str, keep_default_na=False)
    except Exception as e:
        if logger: logger.error(f"pandas Excel fallback failed: {e}")
        confidence.text_extraction = "Failed"
        confidence.overall = "Failed"
        return output

    for sheet_name, df in sheets.items():
        headers = list(df.columns)
        rows = [list(row) for row in df.itertuples(index=False, name=None)]
        col_count = max(len(headers), max((len(r) for r in rows), default=0))
        headers = _pad(headers, col_count)
        rows = [_pad(r, col_count) for r in rows]
        table_md = rows_to_markdown_table(headers, rows)
        output.add_section(body=table_md, heading=f"## {sheet_name}")
        log_info(f"pandas sheet '{sheet_name}' | rows={len(rows)}")

    confidence.text_extraction = "High"
    confidence.table_structure = "Medium"
    confidence.image_extraction = "N/A"
    confidence.image_placement = "N/A"
    confidence.ocr_confidence = "N/A"
    confidence.document_order = "High"
    confidence.derive_overall()

    if progress_callback: progress_callback(1.0)
    return output
